#!/usr/bin/env python
# -*- coding: utf-8 -*-
import pandas as pd
import os
import re
import openpyxl
from openpyxl.styles.borders import Border, Side


class Translation:
    def __init__(self, name=None):
        if name is not None:
            self.load(name)

    def load(self, name):
        ext = os.path.splitext(name)[1].lower()
        if ext == '.xlsx':
            self._df = pd.read_excel(name)
        elif ext == '.json':
            self._df = pd.read_json(name)
        else:
            raise TypeError('not support type: "%s"' % ext)

        self.__process_dataframe()

    def save(self, name, index='English'):
        ext = os.path.splitext(name)[1].lower()
        if ext == '.xlsx':
            self.save_excel(name)
        elif ext == '.json':
            self.save_json(name)
        else:
            raise TypeError('not support type: "%s"' % ext)

    def save_excel(self, name, index='English'):
        def set_styles(ws, frezze_index):
            trans_fill = openpyxl.styles.GradientFill(stop=('D5F5E3', 'EAFAF1'))
            org_fill = openpyxl.styles.GradientFill(stop=('D6EAF8', 'EBF5FB'))
            border = Border(left=Side(style='hair'),
                            right=Side(style='hair'),
                            top=Side(style='hair'),
                            bottom=Side(style='hair'),
                            )
            rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual',
                                                       formula=['""'],
                                                       border=border,
                                                       fill=trans_fill)
            cond_start = chr(ord('A') + frezze_index) + '2'
            cond_end = chr(ord('A') + ws.max_column - 1) + str(ws.max_row)
            ws.conditional_formatting.add('%s:%s' % (cond_start, cond_end), rule)

            start = chr(ord('A') + frezze_index - 1)
            index_cells = ws['%s2:%s%d' % (start, start, ws.max_row)]
            for cell in index_cells:
                cell[0].fill = org_fill
                cell[0].border = border

        frezze_index = self._df.columns.to_list().index(index) + 2
        self._df.to_excel(name, freeze_panes=(1, frezze_index))

        wb = openpyxl.load_workbook(name)
        ws = wb.active
        set_styles(ws, frezze_index)
        font = openpyxl.styles.Font(name='Consolas')
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value is None:
                    cell.value = ''
                elif not isinstance(cell.value, str):
                    cell.value = str(cell.value)
                cell.number_format = '@'
                cell.data_type = 's'
                cell.quotePrefix = True
                cell.font = font
        wb.save(name)

    def save_json(self, name):
        json_str = self._df.to_json(force_ascii=False, indent=4)
        open(name, 'w', encoding='utf-8').write(json_str)

    def __process_dataframe(self):
        self._df.replace(float('nan'), '', inplace=True)
        self._df.drop(columns=filter(lambda x: 'Unnamed' in x or re.search(r'_\d', x), self._df.columns), inplace=True)

    def get_translation(self, index='English'):
        start_index = self._df.columns.to_list().index(index) + 1
        rows = filter(lambda x: sum([len(y) for y in x[start_index:]]) > 0, self._df.itertuples(index=False))
        df = pd.DataFrame(rows)
        df.set_index(index, drop=False, inplace=True)
        for row in df.iterrows():
            for i, cell in enumerate(row[1]):
                if len(cell) == 0:
                    row[1][i] = row[0]
        return df.to_dict('index')

    def set_dataframe(self, df):
        self._df = df
        self.__process_dataframe()

    def set_data(self, data, columns):
        self.set_dataframe(pd.DataFrame(data, columns=columns, dtype=str))

    def get_percent(self, target_index):
        count = len(self._df[self._df[target_index] != ''])
        return count / len(self._df.index) * 100


def try_to_get_translation(name):
    if os.path.exists(name):
        return Translation(name).get_translation()
    else:
        return {}
